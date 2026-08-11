"""Build the NAV history file that the growth chart is drawn from.

The quant feed carries summary returns but no NAV series, so the growth chart
would have nothing to plot. This step fetches the daily NAV of every in-scope
scheme from the public AMFI mirror at api.mfapi.in, keyed on the AMFI scheme code
the feed already carries, and writes one compact file.

Three series end up on the chart, and each is built here:

    fund        the scheme's own daily NAV.
    index       the index itself, from Yahoo through yfinance. These are price
                indices, so they exclude the dividends a fund's NAV already
                contains and read low against the funds by roughly the market's
                dividend yield a year. Where a symbol will not resolve the build
                falls back to an AMFI index tracking scheme, which is dividend
                inclusive but carries that scheme's expense and tracking error.
                Whichever was used is recorded on the series so the chart can say
                which it is showing.
    category    the average of every scored fund in the category. Built from
                daily returns rather than by rebasing NAVs, so a fund that
                launched part way through the window joins the average on the day
                it starts instead of dragging the whole line to its own base.

Resolution is deliberately mixed: every trading day inside the last two years,
one point a week before that. A five year chart is a few hundred pixels wide, so
daily points that far back are bytes nobody can see, and the recent end is where
the shape actually gets read.

Usage:
    python etl/build_navs.py            # incremental, uses the on-disk cache
    python etl/build_navs.py --refresh  # ignore the cache and refetch everything
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from mf import framework as fw  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "data")
CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache", "navs")

API = "https://api.mfapi.in/mf/{code}"

# How far back to keep, and where daily resolution gives way to weekly.
MAX_YEARS = 8
DAILY_DAYS = 730


def _log(msg):
    print(f"[navs] {msg}", flush=True)


# ---------------------------------------------------------------------------
# Fetch
# ---------------------------------------------------------------------------

def fetch_one(code, session, refresh=False):
    """Raw mfapi payload for one scheme code, cached on disk."""
    path = os.path.join(CACHE_DIR, f"{code}.json")
    if not refresh and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except (ValueError, OSError):
            pass  # a truncated cache entry is refetched rather than trusted
    last = None
    for attempt in range(4):
        try:
            r = session.get(API.format(code=code), timeout=45)
            if r.status_code == 200:
                payload = r.json()
                os.makedirs(CACHE_DIR, exist_ok=True)
                with open(path, "w", encoding="utf-8") as fh:
                    json.dump(payload, fh, separators=(",", ":"))
                return payload
            last = f"HTTP {r.status_code}"
        except Exception as e:                      # network, JSON, timeout
            last = str(e)
        time.sleep(2 ** attempt)
    _log(f"  scheme {code} failed: {last}")
    return None


def parse_series(payload):
    """mfapi payload -> [(date, nav)] ascending, cleaned."""
    if not payload or not payload.get("data"):
        return []
    out = []
    for row in payload["data"]:
        try:
            d = dt.datetime.strptime(row["date"], "%d-%m-%Y").date()
            v = float(row["nav"])
        except (KeyError, ValueError, TypeError):
            continue
        if v > 0:
            out.append((d, v))
    out.sort(key=lambda t: t[0])
    return out


def thin(series, today):
    """Daily inside the recent window, weekly before it, capped at MAX_YEARS.

    Weekly points are taken as the last observation in each ISO week, so the
    thinned series still ends each week on a real traded NAV.
    """
    if not series:
        return []
    floor = today - dt.timedelta(days=int(365.25 * MAX_YEARS))
    daily_from = today - dt.timedelta(days=DAILY_DAYS)
    recent, older = [], {}
    for d, v in series:
        if d < floor:
            continue
        if d >= daily_from:
            recent.append((d, v))
        else:
            older[(d.isocalendar()[0], d.isocalendar()[1])] = (d, v)
    out = sorted(older.values()) + recent
    return out


# ---------------------------------------------------------------------------
# Indices
# ---------------------------------------------------------------------------

def _yahoo_session():
    """A requests session Yahoo will actually answer.

    Called cold, Yahoo replies 429 to every request from a fresh client. Asking
    for a cookie first, the way a browser does on the way to a quote page, is
    what makes the difference, so the session is primed once and handed to
    yfinance for every ticker. Without this the whole index fetch fails and the
    build quietly falls back.
    """
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/126.0.0.0 Safari/537.36",
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    })
    for url in ("https://fc.yahoo.com/", "https://finance.yahoo.com/"):
        try:
            s.get(url, timeout=25)
        except Exception:
            pass
    return s


def fetch_index_yahoo(tickers, refresh=False, session=None):
    """First ticker that resolves to a usable daily close series.

    Returns (series, ticker) or (None, None). yfinance is imported lazily so the
    rest of the build still runs where it is not installed.
    """
    try:
        import yfinance as yf
    except ImportError:
        _log("  yfinance is not installed, indices fall back to tracking schemes")
        return None, None

    for t in tickers:
        cache = os.path.join(CACHE_DIR, f"yf_{t.replace('^', '_')}.json")
        if not refresh and os.path.exists(cache):
            try:
                with open(cache, "r", encoding="utf-8") as fh:
                    raw = json.load(fh)
                s = [(dt.date.fromisoformat(d), v) for d, v in raw]
                if len(s) >= 30:
                    _log(f"  {t}: {len(s)} points (cached)")
                    return s, t
            except (ValueError, OSError, TypeError):
                pass
        hist = None
        for period in ("max", "10y"):     # a few symbols reject "max"
            try:
                tk = yf.Ticker(t, session=session) if session else yf.Ticker(t)
                hist = tk.history(period=period, interval="1d", auto_adjust=False)
            except Exception as e:
                _log(f"  {t} [{period}]: {str(e)[:60]}")
                continue
            if hist is not None and not hist.empty and len(hist) > 30:
                break
        if hist is None or hist.empty or "Close" not in hist:
            _log(f"  {t}: no data")
            continue
        s = []
        for idx, close in hist["Close"].items():
            try:
                v = float(close)
            except (TypeError, ValueError):
                continue
            if v > 0:
                s.append((idx.date(), v))
        if len(s) < 30:
            _log(f"  {t}: only {len(s)} points")
            continue
        os.makedirs(CACHE_DIR, exist_ok=True)
        with open(cache, "w", encoding="utf-8") as fh:
            json.dump([[d.isoformat(), v] for d, v in s], fh, separators=(",", ":"))
        _log(f"  {t}: {len(s)} points, {s[0][0]} to {s[-1][0]}")
        return s, t
    return None, None


# ---------------------------------------------------------------------------
# Benchmarks, from the workbook's own benchmark sheet
# ---------------------------------------------------------------------------

def benchmark_series(workbook):
    """Month on month benchmark returns, compounded into index levels.

    The sheet publishes a return per month per benchmark rather than a level, so
    the level is rebuilt by compounding from a base of 100. These are total
    return indices: dividends are already inside them, which is what makes them
    comparable to a fund's NAV rather than reading low against it the way a price
    index does.
    """
    if not workbook or not os.path.exists(workbook):
        _log("  no workbook given, benchmarks skipped")
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        _log("  openpyxl is not installed, benchmarks skipped")
        return {}

    wb = load_workbook(workbook, read_only=True, data_only=True)
    if "BMMom Performance" not in wb.sheetnames:
        _log("  workbook has no BMMom Performance sheet")
        return {}
    it = wb["BMMom Performance"].iter_rows(min_row=4, values_only=True)
    header = list(next(it))
    rows = [r for r in it if r and isinstance(r[0], dt.datetime)]
    rows.sort(key=lambda r: r[0])

    out = {}
    for ci, name in enumerate(header):
        if ci == 0 or not name:
            continue
        name = str(name).strip()
        level, series = 100.0, []
        # The first month's return is the step into it, so the series is seeded a
        # month earlier at the base. Without that the chart would start one month
        # late and lose the first move.
        seeded = False
        for r in rows:
            if ci >= len(r):
                continue
            v = r[ci]
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            d = r[0].date()
            if not seeded:
                prev = (d.replace(day=1) - dt.timedelta(days=1))
                series.append((prev, level))
                seeded = True
            level *= (1.0 + float(v) / 100.0)
            series.append((d, level))
        if len(series) >= 24:
            out[name] = series
    _log(f"  {len(out)} benchmark series, "
         f"{rows[0][0].date()} to {rows[-1][0].date()}")
    return out


# ---------------------------------------------------------------------------
# Category average
# ---------------------------------------------------------------------------

def category_average(series_list):
    """Average daily return across funds, compounded into an index.

    Averaging returns rather than rebased levels means a fund that launched part
    way through the window contributes only from its own start date, instead of
    forcing the whole category line to rebase around it.
    """
    rets = {}          # date -> [return, ...]
    for series in series_list:
        prev = None
        for d, v in series:
            if prev is not None and prev[1] > 0:
                gap = (d - prev[0]).days
                if 0 < gap <= 10:          # skip across long holes in a series
                    rets.setdefault(d, []).append(v / prev[1] - 1.0)
            prev = (d, v)
    if not rets:
        return []
    out, level = [], 100.0
    for d in sorted(rets):
        vals = rets[d]
        if len(vals) < 3:                  # too thin an average to publish
            continue
        level *= (1.0 + sum(vals) / len(vals))
        out.append((d, level))
    return out


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def encode(series):
    return {"d": [d.isoformat() for d, _ in series],
            "v": [round(v, 4) for _, v in series]}


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--refresh", action="store_true",
                    help="ignore the on-disk cache and refetch every scheme")
    ap.add_argument("--workbook",
                    help="Underlying xlsx, read for the benchmark TRI series")
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    with open(os.path.join(DATA_DIR, "funds.json"), "r", encoding="utf-8") as fh:
        funds = json.load(fh)

    in_scope = [f for f in funds
                if f.get("category") in fw.CATEGORIES
                and not fw.excluded_amc(f.get("amc"))
                and f.get("amfiCode")]
    _log(f"{len(in_scope)} in-scope schemes carry an AMFI code")

    # Every index any in-scope category points at, and the tracking scheme that
    # stands in for it if the symbol will not resolve.
    wanted = {fw.index_name_for(c) for c in fw.CATEGORIES}
    index_specs = {n: fw.INDEX_PROXIES[n] for n in wanted}

    jobs = {f["amfiCode"]: f["key"] for f in in_scope}
    for spec in index_specs.values():
        jobs.setdefault(spec["fallback"], None)

    raw = {}
    session = requests.Session()
    session.headers["User-Agent"] = "nse-backend/1.0 (dataset build)"
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futs = {pool.submit(fetch_one, c, session, args.refresh): c for c in jobs}
        done = 0
        for fut in as_completed(futs):
            code = futs[fut]
            raw[code] = fut.result()
            done += 1
            if done % 50 == 0:
                _log(f"  fetched {done}/{len(jobs)}")
    _log(f"fetched {len(jobs)} schemes")

    today = dt.date.today()
    by_key, per_cat = {}, {}
    missing = []
    for f in in_scope:
        s = thin(parse_series(raw.get(f["amfiCode"])), today)
        if len(s) < 30:
            missing.append(f["key"])
            continue
        by_key[f["key"]] = encode(s)
        per_cat.setdefault(f["category"], []).append(s)

    _log("indices:")
    indices = {}
    yf_session = _yahoo_session()
    for name, spec in index_specs.items():
        series, ticker = fetch_index_yahoo(spec["tickers"], args.refresh,
                                           yf_session)
        if series:
            s = thin(series, today)
            indices[name] = {"label": name, "source": "index", "ticker": ticker,
                             "dividends": False, **encode(s)}
            continue
        # No symbol resolved, so stand the tracking scheme in and say so.
        s = thin(parse_series(raw.get(spec["fallback"])), today)
        if len(s) < 30:
            _log(f"  {name}: no index and no tracking scheme, dropped")
            continue
        _log(f"  {name}: falling back to {spec['fallbackLabel']}")
        indices[name] = {"label": spec["fallbackLabel"], "source": "fund",
                         "code": spec["fallback"], "dividends": True, **encode(s)}

    _log("benchmarks:")
    wanted_bm = {fw.benchmark_series_for(c) for c in fw.CATEGORIES}
    benchmarks = {}
    for name, series in benchmark_series(args.workbook).items():
        if name not in wanted_bm:
            continue
        s = [(d, v) for d, v in series
             if d >= today - dt.timedelta(days=int(365.25 * MAX_YEARS))]
        if len(s) >= 24:
            benchmarks[name] = {"label": name, "monthly": True, **encode(s)}
    _log(f"  kept {len(benchmarks)}: {', '.join(sorted(benchmarks)) or 'none'}")

    cat_avg = {}
    for cat, series_list in per_cat.items():
        avg = category_average(series_list)
        if avg:
            cat_avg[cat] = encode(avg)

    payload = {
        "builtAt": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "source": "https://api.mfapi.in",
        "resolution": f"daily for {DAILY_DAYS} days, weekly before that, "
                      f"capped at {MAX_YEARS} years",
        "funds": by_key,
        "indices": indices,
        "indexByCategory": {c: fw.index_name_for(c) for c in fw.CATEGORIES},
        "benchmarks": benchmarks,
        "benchmarkByCategory": {c: fw.benchmark_series_for(c)
                                for c in fw.CATEGORIES},
        "categoryAverage": cat_avg,
    }
    out = os.path.join(DATA_DIR, "navs.json")
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, separators=(",", ":"))

    size = os.path.getsize(out) / 1e6
    _log(f"wrote {out} ({size:.1f} MB)")
    _log(f"  funds with a series : {len(by_key)}")
    _log(f"  indices             : {len(indices)}")
    _log(f"  category averages   : {len(cat_avg)}")
    if missing:
        _log(f"  no usable series    : {len(missing)} ({', '.join(missing[:6])}"
             f"{' ...' if len(missing) > 6 else ''})")


if __name__ == "__main__":
    main()
