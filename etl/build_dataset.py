"""Build the screener dataset.

Two sources, merged on a normalised scheme key:

  1. The Sheety dataPack sheet  -> every quantitative field
  2. The Avendus Automation workbook (optional) -> AUM, manager, holdings

Run:

    python etl/build_dataset.py                          # quant only
    python etl/build_dataset.py --workbook <file.xlsx>   # quant + holdings
    python etl/build_dataset.py --snapshot               # also archive this build

Both arguments are independent: whichever sources are supplied get refreshed, the
rest are left as they are on disk.

The API's column headers are matched by synonym rather than by exact string, so a
header being renamed upstream does not silently drop a metric. Anything that
cannot be matched is printed at the end of the run under "unmapped columns",
which is the one place a quiet data loss would otherwise hide.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import statistics
from collections import defaultdict
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from mf.framework import CATEGORY_MAP, CATEGORIES, benchmark_for  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "data")
HISTORY_DIR = os.path.join(DATA_DIR, "history")

SHEETY_URL = ("https://api.sheety.co/26381234f19b00348c9bb3d7604a8d84"
              "/allFundsQuantData/allFunds")

DEFAULT_BENCHMARK = "Nifty 500 TRI"


# --------------------------------------------------------------------------- #
# column matching
# --------------------------------------------------------------------------- #
#
# The feed publishes flat, single-row headers in a regular shape:
#
#     sharpe [3Y]              ->  sharpe3Y
#     sharpe [3Y]Decile        ->  sharpe3YDecile
#     medianRolling (2017Cutoff) [5Y]  ->  medianRolling2017Cutoff5Y
#     aum (current)            ->  aumCr
#
# So the parser is generic rather than a list of names: pull the trailing
# "Decile", pull the "[horizon]", normalise what is left, and look the stem up.
# A new horizon appearing upstream is picked up without a code change, and
# anything whose stem is unknown is reported instead of silently dropped.

STEMS = {
    "fundname": "fundName",
    "amficode": "amfiCode",
    "category": "category",
    "benchmark": "benchmark",

    "aumcurrentdate": "aumDate",
    "aumcurrent": "aumCr",
    "aum1ydate": "aum1YAgoDate",
    "aum1yago": "aum1YAgoCr",
    "netflow": "netFlowCr",

    "p2p": "return",                    # 1MP2P -> return1M
    "cytd": "returnCYTD",
    "averagecy": "averageCy",
    "cybeatcountvsbenchmark": "cyBeatCount",
    "cybeatvsbenchmark": "cyBeatPct",

    "medianrolling": "medianRolling",
    "medianrolling2017cutoff": "medianRolling2017",

    "sharpe": "sharpe",
    "treynor": "treynor",
    "informationratio": "informationRatio",
    "sortino": "sortino",
    "stddevann": "stdDev",
    "semistddevann": "semiStdDev",
    "maxdrawdown": "maxDrawdown",
    "upsidecapture": "upsideCapture",
    "downsidecapture": "downsideCapture",
    "captureratio": "captureRatio",
    "beta": "beta",

    "largecap": "largeCapPct",
    "midcap": "midCapPct",
    "smallcap": "smallCapPct",
    "cashandothers": "cashPct",
}

# Calendar-year columns (cy24, cy23 ...) are kept as a block rather than as
# individual fields, so a year rolling off does not need a code change.
_CY = re.compile(r"^cy(\d{2})$")
# '3YP2P' and '3YP2PDecile': the horizon leads rather than trailing.
_LEADING_HORIZON = re.compile(r"^(\d+[MY])(.+)$")

TEXT_FIELDS = {"fundName", "category", "benchmark", "fundManager", "amc",
               "aumDate", "aum1YAgoDate", "navDate", "cyBeatCount", "inceptionDate"}


def squash(s):
    return re.sub(r"[^a-z0-9]+", "", str(s or "").lower())


def parse_column(col):
    """(field, is_decile) for one API column, or (None, _) if unrecognised."""
    raw = str(col or "").strip()
    if not raw or raw == "id":
        return None, False

    is_decile = raw.endswith("Decile")
    if is_decile:
        raw = raw[: -len("Decile")]

    horizon = ""
    m = re.search(r"\[([^\]]+)\]\s*$", raw)
    if m:
        horizon = m.group(1).strip().upper().replace(" ", "")
        raw = raw[: m.start()].strip()

    key = squash(raw)

    cy = _CY.match(key)
    if cy:
        field = f"returnCY{cy.group(1)}"
        return (field + "Decile") if is_decile else field, is_decile

    if not horizon:
        lead = _LEADING_HORIZON.match(raw)
        if lead:
            horizon, key = lead.group(1).upper(), squash(lead.group(2))

    stem = STEMS.get(key)
    if stem is None:
        return None, is_decile
    field = stem + horizon if horizon else stem
    return (field + "Decile") if is_decile else field, is_decile


def build_header_map(header_row):
    mapping, unmapped = {}, []
    for col in header_row:
        field, _ = parse_column(col)
        if field is None:
            if str(col).strip() not in ("", "id"):
                unmapped.append(col)
        elif field in mapping.values():
            unmapped.append(col)
        else:
            mapping[col] = field
    return mapping, unmapped


# --------------------------------------------------------------------------- #
# name normalisation - the join key across sources
# --------------------------------------------------------------------------- #

_PLAN_SUFFIX = re.compile(r"\s*-\s*(dir|direct|reg|regular)\b.*$", re.IGNORECASE)
_NOISE_WORDS = re.compile(
    r"\b(fund|funds|scheme|plan|growth|gr|dir|direct|regular|reg|option|"
    r"payout|idcw|dividend\s+reinvestment|and)\b", re.IGNORECASE)

_KEY_ALIASES = {
    "kotakequityopportunities": "kotaklargemidcap",
    "trustflexicap": "trustmfflexicap",
    "trustsmalllcap": "trustmfsmallcap",
    "trustsmallcap": "trustmfsmallcap",
    "trustmidcap": "trustmfmidcap",
    "trustmulticap": "trustmfmulticap",
}


def fund_key(name) -> str:
    s = _PLAN_SUFFIX.sub("", str(name or ""))
    s = _NOISE_WORDS.sub(" ", s)
    s = re.sub(r"[^a-z0-9]+", "", s.lower())
    return _KEY_ALIASES.get(s, s)


def direct_ter(name, value):
    """Expense ratio, but only from a row that is explicitly the direct plan.

    Both plans of a scheme sit in these sheets and only one of them is named. The
    direct plan carries "- Dir"; the regular plan is usually just "Fund Name -
    Growth" with no marker at all, so a filter that drops "- Reg" lets it
    straight through. Since a regular plan runs roughly 0.6 to 1.0 points dearer,
    that quietly published a distributor-loaded fee as the direct one on the
    funds where only the unmarked row existed. Anything not marked direct is
    dropped rather than guessed at.
    """
    if not re.search(r"-\s*Dir\b", str(name), re.I):
        return None
    return num(value)


def clean_display_name(name) -> str:
    return _PLAN_SUFFIX.sub("", str(name or "")).strip()


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        return None if f != f else f
    s = str(v).strip().replace(",", "").replace("%", "")
    if s in ("", "--", "-", "N/A", "#N/A", "NA", "nan", "None", "#DIV/0!", "#VALUE!"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() or None


# --------------------------------------------------------------------------- #
# 1. the quantitative feed
# --------------------------------------------------------------------------- #

def fetch_pack(cache_path=None):
    """Pull the dataPack sheet. Falls back to a local cache when the API is down,
    so a build is never blocked by an upstream outage."""
    import urllib.request

    try:
        req = urllib.request.Request(SHEETY_URL, headers={"User-Agent": "mf-screener/3"})
        with urllib.request.urlopen(req, timeout=180) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        rows = next((v for v in payload.values() if isinstance(v, list)), None)
        if not rows:
            raise ValueError(f"no row array in response: {list(payload)[:5]}")
        if cache_path:
            os.makedirs(os.path.dirname(cache_path), exist_ok=True)
            with open(cache_path, "w", encoding="utf-8") as fh:
                json.dump(payload, fh)
        print(f"  fetched {len(rows)} rows from the dataPack sheet")
        return rows
    except Exception as exc:  # noqa: BLE001 - an offline rebuild is a valid mode
        print(f"  live fetch failed: {exc}")
        if cache_path and os.path.exists(cache_path):
            with open(cache_path, encoding="utf-8") as fh:
                payload = json.load(fh)
            rows = next(v for v in payload.values() if isinstance(v, list))
            print(f"  used cache: {len(rows)} rows")
            return rows
        raise SystemExit(
            f"\nCould not read the quantitative feed and there is no cache to fall "
            f"back on.\n\n  {SHEETY_URL}\n  {exc}\n\n"
            f"Sheety's own error text tells you which end is broken:\n"
            f"  'No project found matching that name'  the project id in the URL is "
            f"wrong or the project was deleted.\n"
            f"  'No sheet found matching that name'    the project is fine, the sheet "
            f"name at the end of the URL is not.\n"
            f"  'Cannot read property ... of undefined' the sheet exists but Sheety "
            f"cannot parse it, which is almost always a blank header row, a blank "
            f"first column, or a merged cell in row 1.\n\n"
            f"Fix the sheet or the URL, then re-run. Nothing was written, so the "
            f"dataset already on disk is untouched.")


def parse_pack(rows):
    """Turn API rows into fund records."""
    if not rows:
        return {}, {}, []
    mapping, unmapped = build_header_map(rows[0].keys())
    print(f"  mapped {len(mapping)} columns; {len(unmapped)} unmapped")

    funds, benchmarks = {}, {}
    for r in rows:
        rec = {}
        for col, field in mapping.items():
            v = r.get(col)
            rec[field] = as_date(v) if field in TEXT_FIELDS else num(v)
        name = (rec.get("fundName") or "").strip()
        if not name:
            continue
        category = (rec.get("category") or "").strip()

        if category.lower().startswith("benchmark"):
            benchmarks[name] = {**rec, "name": name}
            continue

        # Most rows carry a numeric AMFI code. PMS and AIF vehicles carry a
        # sentinel instead, and they are not SEBI mutual fund schemes, so they
        # are kept out of the scored universe rather than ranked against it.
        try:
            amfi_code, vehicle = int(rec.get("amfiCode")), "MF"
        except (TypeError, ValueError):
            amfi_code, vehicle = None, "PMS/AIF"

        rec["key"] = fund_key(name)
        rec["name"] = clean_display_name(name)
        rec["rawName"] = name
        rec["amfiCode"] = amfi_code
        rec["vehicle"] = vehicle
        rec["sourceCategory"] = category
        rec["category"] = (CATEGORY_MAP.get(category.strip().lower())
                           if vehicle == "MF" else None)
        # The feed publishes no benchmark column, so it comes from the category
        # rather than from a blanket default that would read every small cap
        # fund against a broad market index.
        bm, kind = benchmark_for(rec["category"])
        rec["benchmark"] = rec.get("benchmark") or bm
        rec["benchmarkKind"] = kind

        # Net flow as a share of the opening book says more than the rupee figure.
        prev, flow = rec.get("aum1YAgoCr"), rec.get("netFlowCr1Y")
        if prev and flow is not None and prev > 0:
            rec["netFlow1YPct"] = round(100.0 * flow / prev, 1)

        funds[rec["key"]] = rec

    return funds, benchmarks, unmapped


# --------------------------------------------------------------------------- #
# 2. the workbook (AUM, manager, holdings)
# --------------------------------------------------------------------------- #

def parse_workbook(path, funds):
    import openpyxl

    print(f"  opening workbook {os.path.basename(path)} ...")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = set(wb.sheetnames)

    if "Performance" in sheets:
        matched = 0
        for row in wb["Performance"].iter_rows(min_row=5, max_col=13, values_only=True):
            name = row[0]
            if not name or not re.search(r"-\s*Dir\b", str(name)):
                continue
            fund = funds.get(fund_key(name))
            if not fund:
                continue
            matched += 1
            for field, val in (("nav", num(row[1])), ("navDate", as_date(row[2])),
                               ("aumCr", num(row[3])), ("aumDate", as_date(row[4])),
                               ("fundManager", str(row[5]).strip() if row[5] else None),
                               ("return3M", num(row[7])), ("return6M", num(row[8])),
                               ("return1Y", num(row[9])), ("return2Y", num(row[10])),
                               ("return3Y", num(row[11])), ("return5Y", num(row[12]))):
                # The API is authoritative where it carries a value; the workbook
                # only fills the holes.
                if fund.get(field) is None and val is not None:
                    fund[field] = val
        print(f"  Performance: enriched {matched} funds")

    if "Equity Attribute" in sheets:
        matched = 0
        for row in wb["Equity Attribute"].iter_rows(min_row=5, max_col=10, values_only=True):
            name = row[0]
            if not name or re.search(r"-\s*Reg\b", str(name)):
                continue
            fund = funds.get(fund_key(name))
            if not fund:
                continue
            matched += 1
            for field, val in (("amc", str(row[1]).strip() if row[1] else None),
                               ("attrDate", as_date(row[2])), ("ter", direct_ter(name, row[3])),
                               ("largeCapPct", num(row[4])), ("midCapPct", num(row[5])),
                               ("smallCapPct", num(row[6])), ("cashPct", num(row[7]))):
                if fund.get(field) is None and val is not None:
                    fund[field] = val
        print(f"  Equity Attribute: enriched {matched} funds")

    holdings = defaultdict(list)
    if "Underlying Portfolio" in sheets:
        rows_seen = 0
        in_scope = {k for k, f in funds.items() if f.get("category") in CATEGORIES}
        for row in wb["Underlying Portfolio"].iter_rows(min_row=5, max_col=15,
                                                        values_only=True):
            name = row[0]
            if not name:
                continue
            key = fund_key(name)
            if key not in in_scope:
                continue
            pct = num(row[9])
            if pct is None or pct <= 0:
                continue
            rows_seen += 1
            rec = {"s": str(row[1]).strip() if row[1] else "",
                   "sec": str(row[3]).strip() if row[3] else "Unclassified",
                   "pct": round(pct, 3)}
            if row[2]:
                rec["isin"] = str(row[2]).strip()
            if row[4]:
                rec["ins"] = str(row[4]).strip()
            if row[10]:
                rec["mc"] = str(row[10]).strip()
            holdings[key].append(rec)
        for key in holdings:
            holdings[key].sort(key=lambda h: -h["pct"])
        print(f"  Underlying Portfolio: {rows_seen} rows across {len(holdings)} funds")

    # Sheet2 carries the current expense ratio, which the feed does not.
    if "Sheet2" in sheets:
        matched = 0
        for row in wb["Sheet2"].iter_rows(min_row=4, max_col=5, values_only=True):
            name = row[0]
            if not name or re.search(r"-\s*Reg\b", str(name)):
                continue
            fund = funds.get(fund_key(name))
            ter = direct_ter(name, row[1])
            if fund and ter is not None and fund.get("ter") is None:
                fund["ter"] = ter
                matched += 1
        print(f"  Sheet2: expense ratio on {matched} funds")

    # Monthly series: real rolling statistics, and the market cycles the manager
    # block counts against.
    cycles = []
    monthly = parse_monthly(wb)
    if monthly:
        bench_for = {}
        for f in funds.values():
            bm = f.get("benchmark")
            if bm in monthly:
                bench_for[f["key"]] = monthly[bm]
        broad = monthly.get(DEFAULT_BENCHMARK) or monthly.get("Nifty 500 TRI") \
            or monthly.get("BSE 200 TRI")
        cycles = market_cycles(broad)
        print(f"  derived {len(cycles)} market cycles from {DEFAULT_BENCHMARK}: "
              + ", ".join(f"{c['peak']}->{c['trough']} {c['depthPct']}%" for c in cycles))

        enriched = 0
        for raw_name, series in monthly.items():
            key = fund_key(raw_name)
            fund = funds.get(key)
            if not fund or not re.search(r"-\s*Dir\b", str(raw_name)):
                continue
            m = rolling_metrics(series, bench_for.get(key) or broad)
            if not m:
                continue
            # The feed is authoritative for anything it publishes: its rolling
            # medians and ratios span each fund's whole life, while the monthly
            # sheet only reaches back to 2018. These values fill the gaps and
            # supply the one thing the feed does not carry at all, the hit rate.
            for k, v in m.items():
                if fund.get(k) is None:
                    fund[k] = v
            # A live record measured in months beats a bucket inferred from which
            # return horizons happen to be populated.
            if fund.get("vintageYears") is None:
                fund["vintageYears"] = round(len(series) / 12.0, 1)
                fund["vintageBasis"] = f"{len(series)} months of return history"
            enriched += 1
        print(f"  rolling statistics computed for {enriched} funds")

    wb.close()
    return dict(holdings), cycles


# --------------------------------------------------------------------------- #
# 3. monthly return series -> real rolling statistics
# --------------------------------------------------------------------------- #
#
# The workbook carries 97 months of returns per scheme and per benchmark. That is
# enough for 62 rolling three year windows and 38 rolling five year ones, so the
# rolling figures below are computed rather than proxied. Verified against the
# workbook's own stated CAGR: Parag Parikh Flexi Cap 3Y computes to 14.64 against
# a published 14.62, and 5Y to 14.66 against 14.65.

RF_ANNUAL = 7.0          # risk free, per cent per annum
MONTHS_3Y, MONTHS_5Y = 36, 60


def _cagr(monthly):
    """Annualised compound return from a run of monthly per cent returns."""
    p = 1.0
    for r in monthly:
        p *= 1.0 + r / 100.0
    if p <= 0:
        return None
    return (p ** (12.0 / len(monthly)) - 1.0) * 100.0


def _windows(series, size):
    return [series[i:i + size] for i in range(len(series) - size + 1)]


def parse_monthly(wb):
    """Read the two month-on-month sheets into {name: [(date, return %)]}."""
    out = {}
    for sheet in ("Mom Performance", "BMMom Performance"):
        if sheet not in wb.sheetnames:
            continue
        it = wb[sheet].iter_rows(min_row=4, values_only=True)
        header = list(next(it))
        rows = [r for r in it if r and isinstance(r[0], datetime)]
        rows.sort(key=lambda r: r[0])
        for ci, name in enumerate(header):
            if ci == 0 or not name:
                continue
            pts = [(r[0], float(r[ci])) for r in rows
                   if ci < len(r) and isinstance(r[ci], (int, float))
                   and not isinstance(r[ci], bool)]
            if len(pts) >= MONTHS_3Y:
                out[str(name).strip()] = pts
        print(f"  {sheet}: {len(rows)} months, series for "
              f"{sum(1 for k in out)} names so far")
    return out


def rolling_metrics(series, bench):
    """Rolling medians, the rolling hit rate, and the five year ratios.

    The hit rate is the share of rolling three year windows in which the fund beat
    its benchmark. It is the honest version of 'consistency': a fund can carry a
    high median while losing most windows, and this separates the two.
    """
    vals = [v for _, v in series]
    out = {}
    if len(vals) >= MONTHS_3Y:
        w3 = [c for c in (_cagr(w) for w in _windows(vals, MONTHS_3Y)) if c is not None]
        if w3:
            out["medianRolling3Y"] = round(statistics.median(w3), 2)
            out["rollingWindows3Y"] = len(w3)
    if len(vals) >= MONTHS_5Y:
        w5 = [c for c in (_cagr(w) for w in _windows(vals, MONTHS_5Y)) if c is not None]
        if w5:
            out["medianRolling5Y"] = round(statistics.median(w5), 2)

    if bench:
        # Align on date so a fund with a shorter life is compared only over the
        # months both actually lived through.
        bmap = dict(bench)
        pairs = [(d, v, bmap[d]) for d, v in series if d in bmap]
        if len(pairs) >= MONTHS_3Y:
            fv = [p[1] for p in pairs]
            bv = [p[2] for p in pairs]
            wins = 0
            total = 0
            for i in range(len(pairs) - MONTHS_3Y + 1):
                a, b = _cagr(fv[i:i + MONTHS_3Y]), _cagr(bv[i:i + MONTHS_3Y])
                if a is None or b is None:
                    continue
                total += 1
                wins += 1 if a > b else 0
            if total:
                out["rollingHitRate3Y"] = round(100.0 * wins / total, 1)

        if len(pairs) >= MONTHS_5Y:
            fv = [p[1] for p in pairs][-MONTHS_5Y:]
            bv = [p[2] for p in pairs][-MONTHS_5Y:]
            fa, ba = _cagr(fv), _cagr(bv)
            sd = statistics.pstdev(fv) * (12 ** 0.5)
            down = [min(0.0, r - RF_ANNUAL / 12) for r in fv]
            dd = (sum(d * d for d in down) / len(down)) ** 0.5 * (12 ** 0.5)
            te = statistics.pstdev([a - b for a, b in zip(fv, bv)]) * (12 ** 0.5)
            if fa is not None and sd > 0:
                out["sharpe5Y"] = round((fa - RF_ANNUAL) / sd, 4)
            if fa is not None and dd > 0:
                out["sortino5Y"] = round((fa - RF_ANNUAL) / dd, 4)
            if fa is not None and ba is not None and te > 0:
                out["informationRatio5Y"] = round((fa - ba) / te, 4)
    return out


def market_cycles(bench_series, threshold=12.0):
    """Drawdown episodes in a broad benchmark, derived rather than asserted.

    A cycle is a peak to trough fall of at least `threshold` per cent followed by a
    recovery back to the old peak. Using the index's own monthly series means the
    dates come from the data in front of us instead of from a remembered list of
    corrections.
    """
    if not bench_series:
        return []
    level, peak, peak_date = 1.0, 1.0, bench_series[0][0]
    trough, trough_date, in_dd = None, None, False
    cycles = []
    for d, r in bench_series:
        level *= 1.0 + r / 100.0
        if level >= peak:
            if in_dd and trough is not None:
                cycles.append({"peak": peak_date.strftime("%Y-%m"),
                               "trough": trough_date.strftime("%Y-%m"),
                               "recovered": d.strftime("%Y-%m"),
                               "depthPct": round(100.0 * (trough / peak - 1), 1)})
                in_dd = False
            peak, peak_date, trough = level, d, None
            continue
        if trough is None or level < trough:
            trough, trough_date = level, d
        if 100.0 * (1 - level / peak) >= threshold:
            in_dd = True
    if in_dd and trough is not None:
        cycles.append({"peak": peak_date.strftime("%Y-%m"),
                       "trough": trough_date.strftime("%Y-%m"),
                       "recovered": None,
                       "depthPct": round(100.0 * (trough / peak - 1), 1)})
    return cycles


# --------------------------------------------------------------------------- #
# 4. the fund manager master
# --------------------------------------------------------------------------- #

def parse_managers(path, funds, cycles):
    """Per fund manager tenure, experience and cycles run.

    A fund usually has several named managers. Tenure is taken from the longest
    serving one, because the question the block asks is how long this money has
    been run by the people running it now, not how recently somebody joined the
    team. Every named manager is kept for display.
    """
    import openpyxl

    print(f"  opening manager master {os.path.basename(path)} ...")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(min_row=2, max_col=5, values_only=True)

    today = datetime.now(timezone.utc).replace(tzinfo=None)
    per_fund = defaultdict(list)
    unmatched = set()
    for mgr, amc, fund_name, exp, since in it:
        if not fund_name or not mgr:
            continue
        key = fund_key(fund_name)
        if key not in funds:
            unmatched.add(str(fund_name).strip())
            continue
        if isinstance(since, datetime):
            years = (today - since).days / 365.25
            basis = since.strftime("%Y-%m-%d")
        elif str(since or "").strip().lower() == "inception":
            # Managing since inception: tenure is the fund's live record, which we
            # only know once the monthly series has been read. Marked and filled in
            # by the caller rather than guessed here.
            years, basis = None, "inception"
        else:
            years, basis = None, None
        per_fund[key].append({"name": str(mgr).strip(), "amc": str(amc or "").strip(),
                              "experienceYears": num(exp), "sinceBasis": basis,
                              "tenureYears": None if years is None else round(years, 1)})
    wb.close()

    matched, lifted = 0, 0
    for key, mgrs in per_fund.items():
        fund = funds[key]
        # A manager cannot have run a fund before the fund existed, so a dated
        # appointment older than the return window is a hard lower bound on the
        # fund's age. Without this, every fund older than the 97 months the
        # workbook carries is pinned at the same vintage and the track record
        # block stops discriminating for half the universe.
        dated = [m["tenureYears"] for m in mgrs
                 if m["tenureYears"] is not None and m["sinceBasis"] not in (None, "inception")]
        live = num(fund.get("vintageYears"))
        if dated and (live is None or max(dated) > live):
            fund["vintageYears"] = round(max(dated), 1)
            fund["vintageBasis"] = (f"at least {max(dated):.1f}Y: a named manager has run "
                                    f"it since {min(m['sinceBasis'] for m in mgrs if m['sinceBasis'] not in (None, 'inception'))}")
            lifted += 1
        live = num(fund.get("vintageYears"))
        for m in mgrs:
            if m["sinceBasis"] == "inception" and live:
                m["tenureYears"] = round(live, 1)
        tenures = [m["tenureYears"] for m in mgrs if m["tenureYears"] is not None]
        exps = [m["experienceYears"] for m in mgrs if m["experienceYears"] is not None]
        if not tenures:
            continue
        matched += 1
        longest = max(tenures)
        fund["managers"] = mgrs
        fund["managerYears"] = longest
        fund["fundManager"] = ", ".join(m["name"] for m in mgrs)
        if exps:
            fund["managerExperienceYears"] = max(exps)
        # Cycles the longest serving manager has actually run this fund through.
        start = today.timestamp() - longest * 365.25 * 86400
        fund["managerCycles"] = sum(
            1 for c in cycles
            if datetime.strptime(c["trough"], "%Y-%m").timestamp() >= start)

    print(f"  manager master: tenure set on {matched} funds; vintage lifted above the "
          f"return window on {lifted}; {len(unmatched)} scheme names did not match")
    return matched, sorted(unmatched)[:20]


# --------------------------------------------------------------------------- #
# validation
# --------------------------------------------------------------------------- #

def validate(funds, holdings, previous_meta, strict=True):
    """Refuse to write a build that has quietly fallen apart.

    Name-based joins degrade silently: a scheme gets renamed upstream, its key
    stops matching, and the fund simply loses its holdings and its AUM without
    anything failing. These checks turn that into a visible error.
    """
    problems, warnings = [], []
    in_scope = [f for f in funds.values() if f.get("category") in CATEGORIES]

    if not funds:
        problems.append("no funds parsed at all")
    if len(in_scope) < 50:
        problems.append(f"only {len(in_scope)} in-scope funds; expected a few hundred")

    join = 100.0 * len(holdings) / max(1, len(in_scope))
    if holdings and join < 50:
        problems.append(f"holdings joined to only {join:.0f}% of in-scope funds")

    for field, floor in (("medianRolling3Y", 40), ("sortino3Y", 40),
                         ("downsideCapture3Y", 40), ("aumCr", 60)):
        have = 100.0 * sum(1 for f in in_scope if f.get(field) is not None) / max(1, len(in_scope))
        if have < floor:
            warnings.append(f"{field} present on only {have:.0f}% of in-scope funds")

    prev = (previous_meta or {}).get("inScopeFunds")
    if prev and len(in_scope) < prev * 0.8:
        problems.append(f"in-scope count fell from {prev} to {len(in_scope)}, "
                        f"more than a fifth of the universe")

    for w in warnings:
        print(f"  warning: {w}")
    for p in problems:
        print(f"  PROBLEM: {p}")
    if problems and strict:
        raise SystemExit("build refused: the checks above would have written a "
                         "silently broken dataset. Re-run with --force to override.")
    return {"warnings": warnings, "problems": problems}


# --------------------------------------------------------------------------- #
# output
# --------------------------------------------------------------------------- #

def write_json(name, payload):
    os.makedirs(DATA_DIR, exist_ok=True)
    path = os.path.join(DATA_DIR, name)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, separators=(",", ":"))
    print(f"  wrote {name} ({os.path.getsize(path) / 1e6:.2f} MB)")


def snapshot():
    """Archive this build under data/history/<date>/.

    Point in time snapshots cannot be reconstructed after the fact. Every refresh
    that does not archive is a quarter of evidence permanently lost, and without
    the archive there is no way to ever ask whether the model's ranking predicted
    anything.
    """
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    dest = os.path.join(HISTORY_DIR, stamp)
    os.makedirs(dest, exist_ok=True)
    for name in ("funds.json", "meta.json"):
        src = os.path.join(DATA_DIR, name)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(dest, name))
    # Scores, so the archive answers "what did we say at the time" directly.
    try:
        from mf import datastore as ds
        state = ds.load(force=True)
        scores = [{"key": f["key"], "name": f["name"], "category": f["category"],
                   "composite": f.get("composite"), "band": f.get("band"),
                   "categoryRank": f.get("categoryRank"), "tier": f.get("tier"),
                   "evidence": f.get("evidence"),
                   "blocks": f.get("blockScore", {})}
                  for f in state["funds"]]
        with open(os.path.join(dest, "scores.json"), "w", encoding="utf-8") as fh:
            json.dump(scores, fh, separators=(",", ":"))
        print(f"  snapshot: {len(scores)} scored funds archived at data/history/{stamp}/")
    except Exception as exc:  # noqa: BLE001
        print(f"  snapshot: scores not archived ({exc})")


def diff_against_previous(funds):
    """What changed since the last build. This is the monitoring layer: the model
    is recomputed from scratch every time, so movement is only visible if it is
    explicitly diffed."""
    stamps = sorted(os.listdir(HISTORY_DIR)) if os.path.isdir(HISTORY_DIR) else []
    prev_path = None
    for s in reversed(stamps):
        p = os.path.join(HISTORY_DIR, s, "scores.json")
        if os.path.exists(p):
            prev_path = p
            break
    if not prev_path:
        return None
    with open(prev_path, encoding="utf-8") as fh:
        prev = {r["key"]: r for r in json.load(fh)}

    from mf import datastore as ds
    state = ds.load(force=True)
    now = {f["key"]: f for f in state["funds"]}

    added = [now[k]["name"] for k in now if k not in prev]
    dropped = [prev[k]["name"] for k in prev if k not in now]
    moves, band_changes = [], []
    for k in set(now) & set(prev):
        a, b = prev[k].get("composite"), now[k].get("composite")
        if a is not None and b is not None and abs(b - a) >= 5:
            moves.append({"name": now[k]["name"], "from": a, "to": b,
                          "delta": round(b - a, 1)})
        if prev[k].get("band") != now[k].get("band"):
            band_changes.append({"name": now[k]["name"], "from": prev[k].get("band"),
                                 "to": now[k].get("band")})
    moves.sort(key=lambda m: -abs(m["delta"]))
    return {"since": os.path.basename(os.path.dirname(prev_path)),
            "added": added, "dropped": dropped,
            "bandChanges": band_changes, "bigMoves": moves[:25]}


# --------------------------------------------------------------------------- #

def carry_forward(funds):
    """Fill the fresh feed records from the last build wherever the feed is silent.

    The feed and the workbook write into the same record, so re-parsing the feed
    alone leaves every workbook field empty. Anything the new record has no value
    for is taken from the old one; anything the feed does publish wins, because
    an edited sheet is the whole point of the refresh.
    """
    path = os.path.join(DATA_DIR, "funds.json")
    if not os.path.exists(path):
        return 0
    with open(path, encoding="utf-8") as fh:
        old = {f["key"]: f for f in json.load(fh) if f.get("key")}
    filled = 0
    for key, fund in funds.items():
        prev = old.get(key)
        if not prev:
            continue
        for k, v in prev.items():
            if v is not None and fund.get(k) is None:
                fund[k] = v
                filled += 1
    return filled


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", help="Underlying xlsx (AUM, holdings, monthly returns)")
    ap.add_argument("--managers", help="fund manager master xlsx (tenure, experience)")
    ap.add_argument("--snapshot", action="store_true",
                    help="archive this build under data/history/<date>/")
    ap.add_argument("--diff", action="store_true",
                    help="report what changed against the most recent snapshot")
    ap.add_argument("--offline", action="store_true",
                    help="skip the API and rebuild from the cached response")
    ap.add_argument("--force", action="store_true",
                    help="write the dataset even if the validation checks fail")
    args = ap.parse_args()

    cache = os.path.join(ROOT, "etl", "cache", "dataPack.json")
    prev_meta = {}
    meta_path = os.path.join(DATA_DIR, "meta.json")
    if os.path.exists(meta_path):
        with open(meta_path, encoding="utf-8") as fh:
            prev_meta = json.load(fh)

    print("1. quantitative feed")
    if args.offline:
        with open(cache, encoding="utf-8") as fh:
            rows = next(v for v in json.load(fh).values() if isinstance(v, list))
        print(f"  offline: {len(rows)} cached rows")
    else:
        rows = fetch_pack(cache)
    funds, benchmarks, unmapped = parse_pack(rows)
    in_scope = [f for f in funds.values() if f.get("category") in CATEGORIES]
    print(f"  {len(funds)} schemes, {len(in_scope)} inside the eleven equity categories")

    # A run without the workbook is a feed refresh, not a rebuild. The sheet is
    # the authority on which schemes exist and on every number it publishes, but
    # it carries none of what the workbook contributes: holdings, the rolling
    # statistics, cap allocation, expense ratio, manager tenure. Those are
    # carried across from the last build rather than blanked, so the nightly job
    # can pick up an edited sheet without gutting the dataset.
    if not args.workbook:
        carried = carry_forward(funds)
        print(f"  feed refresh: {carried} workbook-derived values carried forward")

    holdings, cycles = {}, []
    if args.workbook:
        print("2. workbook")
        holdings, cycles = parse_workbook(args.workbook, funds)
    else:
        print("2. workbook skipped; holdings left as they are on disk")
        hp = os.path.join(DATA_DIR, "holdings.json")
        if os.path.exists(hp):
            with open(hp, encoding="utf-8") as fh:
                holdings = json.load(fh)

    if args.managers:
        print("3. fund manager master")
        parse_managers(args.managers, funds, cycles)
    else:
        print("3. manager master skipped; the manager block stays unscored")

    print("4. validation")
    checks = validate(funds, holdings, prev_meta, strict=not args.force)

    print("5. writing")
    write_json("funds.json", list(funds.values()))
    write_json("benchmarks.json", benchmarks)
    if args.workbook:
        write_json("holdings.json", holdings)
    write_json("meta.json", {
        "builtAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "totalFunds": len(funds),
        "inScopeFunds": len(in_scope),
        "fundsWithHoldings": len(holdings),
        "categories": {c: sum(1 for f in in_scope if f["category"] == c)
                       for c in CATEGORIES},
        "source": SHEETY_URL,
        "unmappedColumns": unmapped,
        "marketCycles": cycles,
        "checks": checks,
    })

    if unmapped:
        print(f"\n  unmapped columns ({len(unmapped)}), add a synonym if any of these "
              f"should be scored:")
        for c in unmapped:
            print(f"    - {c}")

    if args.snapshot:
        print("6. snapshot")
        snapshot()
    if args.diff:
        d = diff_against_previous(funds)
        print("\n" + (json.dumps(d, indent=2) if d else "  no earlier snapshot to diff against"))

    print("\ndone.")


if __name__ == "__main__":
    main()
